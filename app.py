import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl.styles
from PIL import Image
from datetime import datetime
import io
import numpy as np
import streamlit as st
import tempfile
import os

# Page configuration
st.set_page_config(
    page_title="Energy Consumption Analyzer",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================
# DATA INGESTION FUNCTIONS
# ============================================

def clean_numeric_column(series):
    """Convert European or US number formats to float"""
    if series.dtype == 'object':
        cleaned = series.astype(str).str.replace(',', '.').str.replace(' ', '').str.strip()
        return pd.to_numeric(cleaned, errors='coerce')
    return series

def parse_datetime_robust(date_str, time_str):
    """Try multiple datetime formats (US and EU)"""
    formats_to_try = [
        '%d.%m.%Y %H:%M:%S',
        '%m/%d/%Y %I:%M:%S %p',
        '%m/%d/%Y %H:%M:%S',
        '%d/%m/%Y %H:%M:%S',
        '%Y-%m-%d %H:%M:%S',
        '%d.%m.%Y %H:%M',
        '%m/%d/%Y %H:%M',
    ]
    
    combined = f"{str(date_str).strip()} {str(time_str).strip()}"
    
    for fmt in formats_to_try:
        try:
            return pd.to_datetime(combined, format=fmt)
        except:
            continue
    
    try:
        return pd.to_datetime(combined, infer_datetime_format=True)
    except:
        return pd.NaT

def read_consumption_file(uploaded_file):
    """Read Excel or CSV file from Streamlit uploaded file object"""
    file_ext = uploaded_file.name.split('.')[-1].lower()
    
    if file_ext == 'xlsx':
        try:
            df = pd.read_excel(uploaded_file, sheet_name='Data_RAW', header=None, engine='openpyxl')
        except:
            uploaded_file.seek(0)
            df = pd.read_excel(uploaded_file, header=None, engine='openpyxl')
    
    elif file_ext == 'csv':
        content = uploaded_file.getvalue().decode('utf-8-sig')
        first_line = content.split('\n')[0]
        
        if ';' in first_line:
            delimiter = ';'
        elif ',' in first_line:
            delimiter = ','
        elif '\t' in first_line:
            delimiter = '\t'
        else:
            delimiter = ','
        
        uploaded_file.seek(0)
        
        try:
            df = pd.read_csv(uploaded_file, delimiter=delimiter, header=None, decimal='.', encoding='utf-8-sig')
        except:
            uploaded_file.seek(0)
            try:
                df = pd.read_csv(uploaded_file, delimiter=delimiter, header=None, decimal=',', encoding='utf-8-sig')
            except:
                raise ValueError("Unable to parse CSV file. Check format.")
    
    else:
        raise ValueError(f"Unsupported file format: {file_ext}. Use .xlsx or .csv")
    
    return df

def validate_timeseries(df):
    """Check for data quality issues in time series"""
    issues = []
    
    duplicates = df[df['DATETIME'].duplicated()]
    if len(duplicates) > 0:
        issues.append({
            'type': 'DUPLICATE_TIMESTAMPS',
            'count': len(duplicates),
            'message': f"Found {len(duplicates)} duplicate timestamps",
            'action': 'Will keep first occurrence, remove duplicates'
        })
    
    df_sorted = df.sort_values('DATETIME')
    time_diff = df_sorted['DATETIME'].diff()
    expected_interval = pd.Timedelta(minutes=15)
    
    gaps = time_diff[time_diff > expected_interval * 1.5]
    if len(gaps) > 0:
        total_missing = gaps.sum() / expected_interval
        issues.append({
            'type': 'MISSING_INTERVALS',
            'count': len(gaps),
            'message': f"Found {len(gaps)} gaps in data (approx {int(total_missing)} missing intervals)",
            'action': 'Gaps will be handled during hourly resampling'
        })
    
    non_standard = time_diff[(time_diff != expected_interval) & (time_diff.notna()) & (time_diff < expected_interval * 1.5)]
    if len(non_standard) > 5:
        issues.append({
            'type': 'IRREGULAR_INTERVALS',
            'count': len(non_standard),
            'message': f"Found {len(non_standard)} irregular time intervals",
            'action': 'Will be handled during hourly resampling'
        })
    
    if (df[2] < 0).any():
        neg_count = (df[2] < 0).sum()
        issues.append({
            'type': 'NEGATIVE_VALUES',
            'count': neg_count,
            'message': f"Found {neg_count} negative consumption values",
            'action': 'Negative values will be set to 0'
        })
    
    null_dates = df['DATETIME'].isna().sum()
    null_consumption = df[2].isna().sum()
    if null_dates > 0 or null_consumption > 0:
        issues.append({
            'type': 'NULL_VALUES',
            'count': null_dates + null_consumption,
            'message': f"Found {null_dates} null dates and {null_consumption} null consumption values",
            'action': 'Rows with null values will be removed'
        })
    
    return issues

def show_data_quality_report(issues):
    """Display data quality issues in Streamlit"""
    if not issues:
        st.success("✅ No data quality issues detected! Data is clean and ready for analysis.")
        return True
    
    st.warning("⚠️ Data Quality Issues Detected")
    
    for i, issue in enumerate(issues, 1):
        with st.expander(f"Issue {i}: {issue['type']}", expanded=True):
            st.write(f"**{issue['message']}**")
            st.info(f"Action: {issue['action']}")
    
    st.write("---")
    return True  # Auto-proceed with cleaning

def clean_timeseries(df, issues):
    """Apply cleaning based on identified issues"""
    df_clean = df.copy()
    
    for issue in issues:
        if issue['type'] == 'DUPLICATE_TIMESTAMPS':
            df_clean = df_clean.drop_duplicates(subset='DATETIME', keep='first')
        
        elif issue['type'] == 'NEGATIVE_VALUES':
            df_clean[2] = df_clean[2].clip(lower=0)
        
        elif issue['type'] == 'NULL_VALUES':
            df_clean = df_clean.dropna(subset=['DATETIME', 2])
    
    return df_clean.sort_values('DATETIME').reset_index(drop=True)

def filter_by_date_range(df, filter_mode, selected_year):
    """Filter dataframe based on pre-selected date range"""
    min_date = df['DATETIME'].min()
    max_date = df['DATETIME'].max()
    
    if filter_mode == 'all':
        return df, f"All data ({min_date.date()} to {max_date.date()})"
    
    elif filter_mode == 'recent':
        most_recent_year = df['DATETIME'].dt.year.max()
        filtered = df[df['DATETIME'].dt.year == most_recent_year]
        if len(filtered) == 0:
            return None, f"No data found for {most_recent_year}"
        return filtered, f"Year {most_recent_year} only"
    
    elif filter_mode == 'custom':
        available_years = df['DATETIME'].dt.year.unique()
        if selected_year not in available_years:
            return None, f"No data found for {selected_year}. Available years: {sorted(available_years)}"
        
        filtered = df[df['DATETIME'].dt.year == selected_year]
        if len(filtered) == 0:
            return None, f"No data found for year {selected_year}"
        return filtered, f"Year {selected_year} only"
    
    return df, "All data"

def calculate_load_duration_curve(hourly_df):
    """Generate load duration curve data"""
    consumption = hourly_df['Consumption [kWh]'].values
    sorted_consumption = np.sort(consumption)[::-1]
    hours = np.arange(1, len(sorted_consumption) + 1)
    percentiles = (hours / len(hours)) * 100
    
    return pd.DataFrame({
        'Hours': hours,
        'Consumption [kWh]': sorted_consumption,
        'Percentile': percentiles
    })

# ============================================
# VISUALIZATION FUNCTIONS
# ============================================

def generate_all_plots(hourly_df, monthly_df, raw_df):
    """Generate all visualizations and return as dict of figures"""
    plots = {}
    luminus_green = '#00B612'
    
    df = hourly_df.copy()
    df['Timestamp'] = pd.to_datetime(df['Timestamp'])
    
    # 1. Histogram
    fig1 = plt.figure(figsize=(14, 6))
    sns.histplot(df['Consumption [kWh]'], bins=50, kde=False, color=luminus_green)
    plt.title('Histogram of Hourly Consumption', fontsize=14, fontweight='bold')
    plt.xlabel('Consumption [kWh]')
    plt.ylabel('Frequency')
    plt.grid(axis='y', linestyle='--', alpha=0.3)
    plt.tight_layout()
    plots['histogram'] = fig1
    
    # 2. Time Series
    fig2 = plt.figure(figsize=(14, 6))
    plt.plot(df['Timestamp'], df['Consumption [kWh]'], color=luminus_green, linewidth=0.8)
    plt.title('Hourly Consumption Over Time', fontsize=14, fontweight='bold')
    plt.xlabel('Date')
    plt.ylabel('Consumption [kWh]')
    plt.grid(True, linestyle='--', alpha=0.3)
    plt.gca().xaxis.set_major_locator(mdates.MonthLocator())
    plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%b'))
    plt.tight_layout()
    plots['timeseries'] = fig2
    
    # 3. Daily Profile
    df['Hour'] = df['Timestamp'].dt.hour
    avg_hourly = df.groupby('Hour')['Consumption [kWh]'].mean()
    fig3 = plt.figure(figsize=(14, 6))
    avg_hourly.plot(kind='bar', color=luminus_green, width=0.9)
    plt.title('Average Daily Consumption Profile', fontsize=14, fontweight='bold')
    plt.xlabel('Hour of Day')
    plt.ylabel('Average Consumption [kWh]')
    plt.grid(axis='y', linestyle='--', alpha=0.3)
    plt.tight_layout()
    plots['daily_profile'] = fig3
    
    # 4. Seasonal Patterns
    def get_season(month):
        if month in [12, 1, 2]:
            return 'Winter'
        elif month in [3, 4, 5]:
            return 'Spring'
        elif month in [6, 7, 8]:
            return 'Summer'
        else:
            return 'Fall'
    
    df['Season'] = df['Timestamp'].dt.month.map(get_season)
    seasonal_hourly = df.groupby(['Hour', 'Season'])['Consumption [kWh]'].mean().unstack()
    season_colors = {'Spring': '#DA32A4', 'Summer': '#007BFF', 'Fall': '#FF7F11', 'Winter': '#D72638'}
    
    fig4 = plt.figure(figsize=(14, 6))
    for season in ['Winter', 'Spring', 'Summer', 'Fall']:
        if season in seasonal_hourly.columns:
            plt.plot(seasonal_hourly.index, seasonal_hourly[season], 
                    label=season, color=season_colors[season], linewidth=2)
    plt.title('Average Hourly Consumption by Season', fontsize=14, fontweight='bold')
    plt.xlabel('Hour of Day')
    plt.ylabel('Average Consumption [kWh]')
    plt.grid(True, linestyle='--', alpha=0.3)
    plt.legend()
    plt.tight_layout()
    plots['seasonal'] = fig4
    
    # 5. Monthly Consumption Bar Chart
    df_monthly = monthly_df.copy()
    fig5 = plt.figure(figsize=(14, 6))
    sns.barplot(x='Month', y='Consumption [kWh]', data=df_monthly, color=luminus_green)
    plt.title('Monthly Energy Consumption', fontsize=14, fontweight='bold')
    plt.xlabel('Month')
    plt.ylabel('Consumption [kWh]')
    plt.xticks(rotation=45)
    plt.grid(axis='y', linestyle='--', alpha=0.3)
    plt.tight_layout()
    plots['monthly_bar'] = fig5
    
    # 6. Capacity Tariff - Top 10 Peaks
    df_raw_analysis = raw_df.copy()
    df_raw_analysis = df_raw_analysis.rename(columns={'Consumption': 'Consumption [kWh]'})
    
    top10 = df_raw_analysis.nlargest(10, 'Consumption [kWh]').sort_values('Timestamp')
    
    fig6 = plt.figure(figsize=(14, 6))
    plt.plot(df_raw_analysis['Timestamp'], df_raw_analysis['Consumption [kWh]'], 
            color=luminus_green, linewidth=0.8, label='Consumption', alpha=0.7)
    plt.scatter(top10['Timestamp'], top10['Consumption [kWh]'], 
               color='red', s=100, label='Top 10 Peaks', zorder=5)
    plt.title('Time Series with Top 10 Peaks', fontsize=14, fontweight='bold')
    plt.xlabel('Date')
    plt.ylabel('Consumption [kWh]')
    plt.grid(True, linestyle='--', alpha=0.3)
    plt.gca().xaxis.set_major_locator(mdates.MonthLocator())
    plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%b'))
    plt.legend()
    plt.tight_layout()
    plots['top10_peaks'] = fig6
    
    # 7. Top 100 by Season
    top100 = df_raw_analysis.nlargest(100, 'Consumption [kWh]').copy()
    top100['Season'] = top100['Timestamp'].dt.month.map(get_season)
    season_counts = top100['Season'].value_counts().reindex(['Spring', 'Summer', 'Fall', 'Winter'])
    
    fig7 = plt.figure(figsize=(10, 6))
    season_counts.plot(kind='bar', color=[season_colors[s] for s in season_counts.index])
    plt.title('Top 100 Consumption Values by Season', fontsize=14, fontweight='bold')
    plt.xlabel('Season')
    plt.ylabel('Count')
    plt.xticks(rotation=0)
    plt.grid(axis='y', linestyle='--', alpha=0.3)
    plt.tight_layout()
    plots['top100_seasons'] = fig7
    
    # 8. Load Duration Curve
    ldc_data = calculate_load_duration_curve(hourly_df)
    fig8 = plt.figure(figsize=(14, 6))
    plt.plot(ldc_data['Percentile'], ldc_data['Consumption [kWh]'], 
            color=luminus_green, linewidth=2)
    plt.axhline(y=hourly_df['Consumption [kWh]'].mean(), color='red', 
               linestyle='--', label=f"Average ({hourly_df['Consumption [kWh]'].mean():.1f} kWh)")
    plt.axhline(y=hourly_df['Consumption [kWh]'].median(), color='orange', 
               linestyle='--', label=f"Median ({hourly_df['Consumption [kWh]'].median():.1f} kWh)")
    plt.title('Load Duration Curve', fontsize=14, fontweight='bold')
    plt.xlabel('Percentage of Time (%)')
    plt.ylabel('Consumption [kWh]')
    plt.grid(True, linestyle='--', alpha=0.3)
    plt.legend()
    plt.tight_layout()
    plots['load_duration'] = fig8
    
    # Store additional data for later use
    plots['seasonal_hourly_data'] = seasonal_hourly
    plots['top10_data'] = top10
    plots['top100_data'] = top100
    plots['ldc_data'] = ldc_data
    
    return plots

def create_excel_file(hourly_df, monthly_df, raw_df, plots, filename):
    """Create Excel file in memory with all data and visualizations"""
    
    # Prepare data
    hourly_df_no_feb29 = hourly_df[~((hourly_df['Timestamp'].dt.month == 2) & 
                                     (hourly_df['Timestamp'].dt.day == 29))].reset_index(drop=True)
    
    raw_df_for_save = raw_df[['Date', 'Time', 'Consumption']].copy()
    
    # Create temporary directory for images
    with tempfile.TemporaryDirectory() as tmpdir:
        # Save plots as images
        plot_files = {}
        
        for name, fig in plots.items():
            if name.endswith('_data'):
                continue
            filepath = os.path.join(tmpdir, f'{name}.png')
            fig.savefig(filepath, dpi=100, bbox_inches='tight')
            plot_files[name] = filepath
        
        # Create Excel file in memory
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            hourly_df.to_excel(writer, sheet_name='Hourly', index=False)
            monthly_df.to_excel(writer, sheet_name='Monthly', index=False)
            hourly_df_no_feb29.to_excel(writer, sheet_name='Sympheny', index=False)
            raw_df_for_save.to_excel(writer, sheet_name='Data_RAW', index=False, header=False)
        
        # Load workbook to add images
        output.seek(0)
        wb = load_workbook(output)
        
        # Add visualizations to sheets
        def add_image_to_sheet(sheet_name, image_path, anchor='A1'):
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(title=sheet_name)
            else:
                ws = wb[sheet_name]
            img = XLImage(image_path)
            img.anchor = anchor
            ws.add_image(img)
        
        # Add histogram
        add_image_to_sheet('Histogram', plot_files['histogram'])
        
        # Add time series
        add_image_to_sheet('Time Series Yearly', plot_files['timeseries'])
        
        # Add daily profile
        add_image_to_sheet('Daily Profile', plot_files['daily_profile'])
        
        # Add seasonal patterns with data
        ws_seasonal = wb.create_sheet(title='Weekly patterns')
        seasonal_data = plots['seasonal_hourly_data'].reset_index()
        for r_idx, row in enumerate(dataframe_to_rows(seasonal_data, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_seasonal.cell(row=r_idx, column=c_idx, value=value)
        img = XLImage(plot_files['seasonal'])
        img.anchor = 'F1'
        ws_seasonal.add_image(img)
        
        # Add monthly chart to Monthly sheet
        ws_monthly = wb['Monthly']
        img = XLImage(plot_files['monthly_bar'])
        img.anchor = 'F1'
        ws_monthly.add_image(img)
        
        # Add capacity tariff analysis
        ws_top10 = wb.create_sheet(title='Quarter top 10')
        top10_data = plots['top10_data'][['Timestamp', 'Consumption [kWh]']]
        for r_idx, row in enumerate(dataframe_to_rows(top10_data, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_top10.cell(row=r_idx, column=c_idx, value=value)
        img = XLImage(plot_files['top10_peaks'])
        img.anchor = 'E1'
        ws_top10.add_image(img)
        
        ws_top100 = wb.create_sheet(title='Quarter top 100')
        top100_data = plots['top100_data'][['Timestamp', 'Consumption [kWh]', 'Season']]
        for r_idx, row in enumerate(dataframe_to_rows(top100_data, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_top100.cell(row=r_idx, column=c_idx, value=value)
        img = XLImage(plot_files['top100_seasons'])
        img.anchor = 'E1'
        ws_top100.add_image(img)
        
        # Add Load Duration Curve
        ws_ldc = wb.create_sheet(title='Load Duration Curve')
        ldc_data = plots['ldc_data']
        for r_idx, row in enumerate(dataframe_to_rows(ldc_data, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_ldc.cell(row=r_idx, column=c_idx, value=value)
        img = XLImage(plot_files['load_duration'])
        img.anchor = 'E1'
        ws_ldc.add_image(img)
        
        # Save to BytesIO
        final_output = io.BytesIO()
        wb.save(final_output)
        final_output.seek(0)
        
        return final_output

# ============================================
# MAIN PROCESSING FUNCTION
# ============================================

@st.cache_data(show_spinner=False)
def process_data_streamlit(uploaded_file_content, file_name, filter_mode, selected_year):
    """Main processing function adapted for Streamlit"""
    
    # Create a file-like object from bytes
    uploaded_file = io.BytesIO(uploaded_file_content)
    uploaded_file.name = file_name
    
    # Stage 1: Load data
    df = read_consumption_file(uploaded_file)
    
    # Data cleaning
    df[2] = clean_numeric_column(df[2])
    df['DATETIME'] = df.apply(
        lambda row: parse_datetime_robust(row[0], row[1]), axis=1
    )
    
    # Stage 2: Validate
    issues = validate_timeseries(df)
    
    # Clean data
    df = clean_timeseries(df, issues)
    
    # Apply date filter
    df, date_range_info = filter_by_date_range(df, filter_mode, selected_year)
    
    if df is None or len(df) == 0:
        return None, None, f"No data to process. {date_range_info}"
    
    # Stage 3: Process data
    raw_df = df[[0, 1, 2, 'DATETIME']].copy()
    raw_df.columns = ['Date', 'Time', 'Consumption', 'Timestamp']
    
    df.set_index('DATETIME', inplace=True)
    hourly_df = df[2].resample('H').sum().reset_index()
    hourly_df[2] = hourly_df[2] / 4
    hourly_df.columns = ['Timestamp', 'Consumption [kWh]']
    
    monthly_df = hourly_df.copy()
    monthly_df['Month'] = monthly_df['Timestamp'].dt.to_period('M')
    monthly_df = monthly_df.groupby('Month')['Consumption [kWh]'].sum().reset_index()
    monthly_df['Month'] = monthly_df['Month'].astype(str)
    
    # Stage 4: Generate visualizations
    plots = generate_all_plots(hourly_df, monthly_df, raw_df)
    
    # Stage 5: Create Excel file
    excel_file = create_excel_file(hourly_df, monthly_df, raw_df, plots, file_name)
    
    return {
        'hourly_df': hourly_df,
        'monthly_df': monthly_df,
        'raw_df': raw_df,
        'plots': plots,
        'excel_file': excel_file,
        'date_range_info': date_range_info,
        'issues': issues
    }, None, None

# ============================================
# STREAMLIT UI
# ============================================

def display_visualizations(plots):
    """Display all plots in organized layout"""
    
    st.subheader("📊 Consumption Overview")
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("**Histogram of Hourly Consumption**")
        st.pyplot(plots['histogram'])
    
    with col2:
        st.markdown("**Time Series - Yearly View**")
        st.pyplot(plots['timeseries'])
    
    st.divider()
    
    st.subheader("📈 Consumption Patterns")
    col3, col4 = st.columns(2)
    
    with col3:
        st.markdown("**Average Daily Profile**")
        st.pyplot(plots['daily_profile'])
    
    with col4:
        st.markdown("**Seasonal Hourly Patterns**")
        st.pyplot(plots['seasonal'])
    
    st.divider()
    
    st.subheader("📅 Monthly Analysis")
    st.pyplot(plots['monthly_bar'])
    
    st.divider()
    
    st.subheader("⚡ Peak Demand Analysis")
    col5, col6 = st.columns(2)
    
    with col5:
        st.markdown("**Top 10 Consumption Peaks**")
        st.pyplot(plots['top10_peaks'])
    
    with col6:
        st.markdown("**Top 100 Peaks by Season**")
        st.pyplot(plots['top100_seasons'])
    
    st.divider()
    
    st.subheader("📉 Load Duration Curve")
    st.pyplot(plots['load_duration'])

def display_data_tables(results):
    """Display data tables"""
    
    tab1, tab2, tab3, tab4 = st.tabs([
        "Hourly Data", 
        "Monthly Summary", 
        "Top 10 Peaks",
        "Load Duration Curve"
    ])
    
    with tab1:
        st.dataframe(
            results['hourly_df'].style.format({'Consumption [kWh]': '{:.2f}'}),
            use_container_width=True,
            height=400
        )
        st.caption(f"Total rows: {len(results['hourly_df'])}")
    
    with tab2:
        st.dataframe(
            results['monthly_df'].style.format({'Consumption [kWh]': '{:.2f}'}),
            use_container_width=True
        )
        total_consumption = results['monthly_df']['Consumption [kWh]'].sum()
        st.metric("Total Consumption", f"{total_consumption:,.2f} kWh")
    
    with tab3:
        top10_display = results['plots']['top10_data'][['Timestamp', 'Consumption [kWh]']].copy()
        st.dataframe(
            top10_display.style.format({'Consumption [kWh]': '{:.2f}'}),
            use_container_width=True
        )
    
    with tab4:
        st.dataframe(
            results['plots']['ldc_data'].style.format({
                'Consumption [kWh]': '{:.2f}',
                'Percentile': '{:.2f}'
            }),
            use_container_width=True,
            height=400
        )

def main():
    """Main Streamlit application"""
    
    # Custom CSS
    st.markdown("""
        <style>
        .main > div {
            padding-top: 2rem;
        }
        .stAlert {
            margin-top: 1rem;
            margin-bottom: 1rem;
        }
        </style>
    """, unsafe_allow_html=True)
    
    # Header
    st.title("⚡ Energy Consumption Analyzer")
    st.markdown("Automated analysis of quarter-hourly energy data")
    
    # Sidebar for inputs
    with st.sidebar:
        st.header("📁 Data Source")
        uploaded_file = st.file_uploader(
            "Upload your data file",
            type=['xlsx', 'csv'],
            help="Supports Excel (.xlsx) and CSV files with quarter-hourly consumption data"
        )
        
        st.info("✓ Handles US and EU date/number formats automatically")
        
        st.divider()
        
        st.header("📅 Date Range")
        date_selection = st.radio(
            "Select date range:",
            options=['all', 'recent', 'custom'],
            format_func=lambda x: {
                'all': '📊 Analyze all available data',
                'recent': '📈 Most recent full year only',
                'custom': '🎯 Specific year'
            }[x],
            index=0
        )
        
        selected_year = None
        if date_selection == 'custom':
            current_year = datetime.now().year
            selected_year = st.selectbox(
                "Select year:",
                options=list(range(current_year, 2019, -1)),
                index=0
            )
        else:
            selected_year = datetime.now().year
        
        st.divider()
        
        analyze_button = st.button(
            "🚀 Start Analysis",
            type="primary",
            disabled=uploaded_file is None,
            use_container_width=True
        )
    
    # Main content area
    if uploaded_file is None:
        # Welcome screen
        st.info("👈 Upload a file to begin analysis")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("### ℹ️ Expected Data Format")
            st.markdown("""
            Your file should contain quarter-hourly consumption data:
            - **Column 0**: Date
            - **Column 1**: Time  
            - **Column 2**: Consumption value
            
            **Supported formats:**
            - Date: DD.MM.YYYY, MM/DD/YYYY, YYYY-MM-DD
            - Time: HH:MM:SS, HH:MM
            - Numbers: Both US (.) and EU (,) decimal formats
            """)
        
        with col2:
            st.markdown("### 📋 Features")
            st.markdown("""
            This tool automatically generates:
            - ✅ Data quality validation
            - 📊 Multiple visualization types
            - 📈 Hourly and monthly summaries
            - ⚡ Peak demand analysis
            - 📉 Load duration curves
            - 📥 Comprehensive Excel report
            """)
    
    elif analyze_button:
        # Process the data
        with st.spinner("🔄 Processing your data..."):
            # Read file content once
            file_content = uploaded_file.getvalue()
            file_name = uploaded_file.name
            
            results, _, error = process_data_streamlit(
                file_content,
                file_name,
                date_selection, 
                selected_year
            )
        
        if error:
            st.error(f"❌ {error}")
            st.stop()
        
        if results is None:
            st.error("❌ Failed to process data. Please check your file format.")
            st.stop()
        
        # Show data quality report
        with st.expander("🔍 Data Quality Report", expanded=bool(results['issues'])):
            show_data_quality_report(results['issues'])
        
        st.success(f"✅ Analysis complete! {results['date_range_info']}")
        
        # Display tabs with results
        tab1, tab2, tab3 = st.tabs([
            "📊 Visualizations",
            "📋 Data Tables",
            "⬇️ Download"
        ])
        
        with tab1:
            display_visualizations(results['plots'])
        
        with tab2:
            display_data_tables(results)
        
        with tab3:
            st.markdown("### 📥 Download Complete Report")
            st.markdown("Download an Excel file containing all data, visualizations, and analysis.")
            
            timestamp_str = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"energy_analysis_{timestamp_str}.xlsx"
            
            st.download_button(
                label="📥 Download Excel Report",
                data=results['excel_file'],
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            
            st.info("💡 The Excel file includes multiple sheets with hourly data, monthly summaries, visualizations, and detailed analysis.")

if __name__ == "__main__":
    main()